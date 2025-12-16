"""Inventory service for managing inventory uploads and items."""
from typing import List, Dict, Any, Optional
from werkzeug.datastructures import FileStorage
import os
import io
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from app.extensions import db
from app.models import InventoryUpload, InventoryItem
from app.services.excel_service import ExcelService, ExcelParsingError


class InventoryService:
    """Service for inventory operations."""

    @staticmethod
    def process_upload(file: FileStorage,
                      upload_name: Optional[str] = None,
                      site: Optional[str] = None,
                      site2: Optional[str] = None,
                      bsf_factor: float = 0.63) -> InventoryUpload:
        """Process an inventory file upload.

        Args:
            file: Uploaded file
            upload_name: Name for the upload
            site: Primary site
            site2: Secondary site
            bsf_factor: Space utilization factor (default 0.63)

        Returns:
            Created InventoryUpload object

        Raises:
            ExcelParsingError: If file cannot be parsed
            ValueError: If file is invalid
        """
        # Validate file
        if not file or file.filename == '':
            raise ValueError("No file provided")

        # Check file extension
        allowed_extensions = {'xlsx', 'xls'}
        file_ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
        if file_ext not in allowed_extensions:
            raise ValueError(f"Invalid file type. Allowed types: {', '.join(allowed_extensions)}")

        # Save file temporarily (cross-platform)
        temp_dir = tempfile.gettempdir()
        temp_path = os.path.join(temp_dir, file.filename)
        file.save(temp_path)

        try:
            # Parse Excel file
            excel_service = ExcelService()
            parsed_data = excel_service.parse_inventory_file(temp_path)

            # Extract and normalize items
            normalized_items = excel_service.extract_inventory_items(
                parsed_data['items'],
                site=site,
                site2=site2
            )

            # Calculate summary statistics
            summary_stats = excel_service.calculate_summary_stats(normalized_items)

            # Create inventory upload record
            inventory_upload = InventoryUpload(
                upload_name=upload_name or file.filename,
                filename=file.filename,
                site=site,
                site2=site2,
                total_items=summary_stats['total_items'],
                total_entries=summary_stats['total_entries'],
                total_weight=summary_stats['total_weight'],
                total_area=summary_stats['total_area'],
                bsf_factor=bsf_factor,
                upload_metadata={
                    'categories': summary_stats['categories'],
                    'columns': parsed_data['metadata']['columns']
                }
            )

            db.session.add(inventory_upload)
            db.session.flush()  # Get the upload ID

            # Create inventory items
            for item_data in normalized_items:
                inventory_item = InventoryItem(
                    upload_id=inventory_upload.id,
                    name=item_data['name'],
                    description=item_data.get('description'),
                    quantity=item_data.get('quantity', 1),
                    category=item_data.get('category'),
                    weight=item_data.get('weight'),
                    length=item_data.get('length'),
                    width=item_data.get('width'),
                    height=item_data.get('height'),
                    area=item_data.get('area'),
                    psf=item_data.get('psf'),
                    service_branch=item_data.get('service_branch'),
                    item_data=item_data.get('item_data')
                )

                # Calculate area if needed
                inventory_item.calculate_area()
                # Calculate PSF if needed
                inventory_item.calculate_psf()

                db.session.add(inventory_item)

            db.session.commit()

            return inventory_upload

        finally:
            # Clean up temporary file
            if os.path.exists(temp_path):
                os.remove(temp_path)

    @staticmethod
    def get_upload_with_items(upload_id: int) -> Optional[InventoryUpload]:
        """Get an inventory upload with its items.

        Args:
            upload_id: Upload ID

        Returns:
            InventoryUpload object or None
        """
        return InventoryUpload.query.get(upload_id)

    @staticmethod
    def get_upload_items(upload_id: int,
                        category: Optional[str] = None,
                        limit: Optional[int] = None,
                        offset: int = 0) -> List[InventoryItem]:
        """Get items for an upload with optional filtering.

        Args:
            upload_id: Upload ID
            category: Filter by category
            limit: Limit number of results
            offset: Offset for pagination

        Returns:
            List of InventoryItem objects
        """
        query = InventoryItem.query.filter_by(upload_id=upload_id)

        if category:
            query = query.filter_by(category=category)

        query = query.offset(offset)

        if limit:
            query = query.limit(limit)

        return query.all()

    @staticmethod
    def update_bsf_factor(upload_id: int, bsf_factor: float) -> InventoryUpload:
        """Update the BSF factor for an upload.

        Args:
            upload_id: Upload ID
            bsf_factor: New BSF factor (0.0 - 1.0)

        Returns:
            Updated InventoryUpload object

        Raises:
            ValueError: If BSF factor is invalid
        """
        if not 0.0 <= bsf_factor <= 1.0:
            raise ValueError("BSF factor must be between 0.0 and 1.0")

        upload = InventoryUpload.query.get_or_404(upload_id)
        upload.bsf_factor = bsf_factor
        db.session.commit()

        return upload

    @staticmethod
    def delete_upload(upload_id: int) -> None:
        """Delete an inventory upload and all its items.

        Args:
            upload_id: Upload ID
        """
        upload = InventoryUpload.query.get_or_404(upload_id)
        db.session.delete(upload)
        db.session.commit()

    @staticmethod
    def get_summary_statistics(upload_id: int) -> Dict[str, Any]:
        """Get summary statistics for an upload.

        Args:
            upload_id: Upload ID

        Returns:
            Dictionary of summary statistics
        """
        upload = InventoryUpload.query.get_or_404(upload_id)
        items = upload.items.all()

        # Category breakdown
        category_stats = {}
        service_branch_stats = {}

        for item in items:
            # Category stats
            cat = item.category or 'Uncategorized'
            if cat not in category_stats:
                category_stats[cat] = {
                    'count': 0,
                    'total_quantity': 0,
                    'total_weight': 0,
                    'total_area': 0
                }
            category_stats[cat]['count'] += 1
            category_stats[cat]['total_quantity'] += item.quantity
            category_stats[cat]['total_weight'] += (item.weight or 0) * item.quantity
            category_stats[cat]['total_area'] += (item.area or 0) * item.quantity

            # Service branch stats
            branch = item.service_branch or 'Unassigned'
            if branch not in service_branch_stats:
                service_branch_stats[branch] = {
                    'count': 0,
                    'total_quantity': 0
                }
            service_branch_stats[branch]['count'] += 1
            service_branch_stats[branch]['total_quantity'] += item.quantity

        # Items requiring special handling
        climate_controlled_count = upload.items.filter_by(requires_climate_control=True).count()
        special_handling_count = upload.items.filter_by(requires_special_handling=True).count()

        return {
            'upload_id': upload.id,
            'upload_name': upload.upload_name,
            'total_items': upload.total_items,
            'total_entries': upload.total_entries,
            'total_weight': float(upload.total_weight) if upload.total_weight else 0,
            'total_area': float(upload.total_area) if upload.total_area else 0,
            'bsf_factor': float(upload.bsf_factor) if upload.bsf_factor else 0.63,
            'categories': category_stats,
            'service_branches': service_branch_stats,
            'climate_controlled_items': climate_controlled_count,
            'special_handling_items': special_handling_count
        }

    @staticmethod
    def export_to_excel(upload_id: int) -> io.BytesIO:
        """Export inventory upload to Excel file.

        Args:
            upload_id: Upload ID

        Returns:
            BytesIO buffer containing Excel file

        Raises:
            NotFound: If upload doesn't exist
        """
        upload = InventoryUpload.query.get_or_404(upload_id)
        items = upload.items.all()

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"

        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Define headers
        headers = [
            'Name', 'Description', 'Quantity', 'Category',
            'Weight (lbs)', 'Length (ft)', 'Width (ft)', 'Height (ft)',
            'Area (sq ft)', 'PSF', 'Service Branch',
            'Climate Control', 'Special Handling'
        ]

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Write data
        for row_num, item in enumerate(items, 2):
            ws.cell(row=row_num, column=1, value=item.name)
            ws.cell(row=row_num, column=2, value=item.description or '')
            ws.cell(row=row_num, column=3, value=item.quantity)
            ws.cell(row=row_num, column=4, value=item.category or '')
            ws.cell(row=row_num, column=5, value=float(item.weight) if item.weight else None)
            ws.cell(row=row_num, column=6, value=float(item.length) if item.length else None)
            ws.cell(row=row_num, column=7, value=float(item.width) if item.width else None)
            ws.cell(row=row_num, column=8, value=float(item.height) if item.height else None)
            ws.cell(row=row_num, column=9, value=float(item.area) if item.area else None)
            ws.cell(row=row_num, column=10, value=float(item.psf) if item.psf else None)
            ws.cell(row=row_num, column=11, value=item.service_branch or '')
            ws.cell(row=row_num, column=12, value='Yes' if item.requires_climate_control else 'No')
            ws.cell(row=row_num, column=13, value='Yes' if item.requires_special_handling else 'No')

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer
